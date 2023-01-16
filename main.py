# Read Spreadsheet File From Same Directory

# Task:

# 列出每家公司有多少商品
# 列出存貨少於10的商品
# 列出每個公司的總存貨數量
# 列出每間公司的總存庫價值
# 計算每間公司每個產品的存庫價值，並且入excel表

import openpyxl
from copy import copy

spreadsheet_file_path = input("enter spreadsheet file path.\n")

try:
    inv_file = openpyxl.load_workbook(spreadsheet_file_path)
except openpyxl.utils.exceptions.InvalidFileException:
    raise Exception(f'Can not find this file')


product_list = inv_file["Sheet1"]

products_per_supplier = {}
products_under_10_inv = {}
inventory_per_company = {}
total_value_per_supplier = {}
newValueCell = product_list.cell(1, 5, "Value")

if product_list.cell(1, 1).has_style:
    newValueCell._style = copy(product_list.cell(1, 1)._style)

for product_row in range(2, product_list.max_row + 1):
    company_name = product_list.cell(product_row, 4).value
    price = product_list.cell(product_row, 3).value
    inventory = product_list.cell(product_row, 2).value
    product_no = int(product_list.cell(product_row, 1).value)
    value = price * inventory
    # 計算每間公司每個產品的存庫價值，並且入excel表
    inventory_value_cell = product_list.cell(product_row, 5)
    inventory_value_cell.value = value

    # 列出每家公司有多少商品
    if company_name in products_per_supplier:
        products_per_supplier[company_name] += 1
    else:
        products_per_supplier[company_name] = 1

    # 列出每個公司的總存貨數量
    if company_name in inventory_per_company:
        inventory_per_company[company_name] = int(inventory_per_company[company_name] + inventory)
    else:
        inventory_per_company[company_name] = int(inventory)

    # 計算每間公司每個產品的存庫價值，並且入excel表
    if company_name in total_value_per_supplier:
        total_value_per_supplier[company_name] += value
    else:
        total_value_per_supplier[company_name] = value

    # 列出存貨少於10的商品
    if inventory < 10:
        products_under_10_inv[product_no] = int(inventory)

    # 列出每間公司的總存庫價值
    if company_name in total_value_per_supplier:
        total_value_per_supplier[company_name] += value
    else:
        total_value_per_supplier[company_name] = value

print(f"列出每家公司有多少商品 {products_per_supplier}")
print(f"列出存貨少於10的商品 {products_under_10_inv}")
print(f"列出每個公司的總存貨數量 {inventory_per_company}")
print(f"列出每間公司的總存庫價值 {total_value_per_supplier}")

# save to new file
inv_file.save(f"with_total_value_{spreadsheet_file_path}")
